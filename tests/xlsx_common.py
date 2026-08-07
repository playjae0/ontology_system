# -*- coding: utf-8 -*-
"""raw mock 생성 공통 유틸 — 헤더·서식·연속값 접기(병합/상동)"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HDR_FONT = Font(bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14)
BODY = Font(size=10)
THIN = Side(style="thin", color="999999")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
DATA_START = 4          # 1행 제목 · 2행 문서정보 · 3행 헤더 (header_row=3)


def new_sheet(title, sheet_name, doc_meta, headers, widths, fill="D9E1F2"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A2"] = doc_meta
    ws["A2"].font = Font(size=9, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    hf = PatternFill("solid", fgColor=fill)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = HDR_FONT, hf, CEN, BOX
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 28
    return wb, ws


def put_rows(ws, rows, align_cols=(), start=DATA_START):
    for r, row in enumerate(rows, start=start):
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font, c.border = BODY, BOX
            c.alignment = CEN if i in align_cols else LEFT
        ws.row_dimensions[r].height = 22


def collapse(ws, col, values, mode="merge", start=DATA_START, keys=None):
    """연속 동일값 구간을 접는다.
    mode="merge" → 세로 병합(종속 칸은 빈 셀)   / normalizer의 병합 전개 재료
    mode="ditto" → 둘째 칸부터 상동 기호 '〃'   / normalizer의 상동 해소 재료
    keys       → 지정 시 이 값이 바뀌면 구간을 끊는다(상위 그룹 밖으로 병합이 번지는 것 방지)
    반환: 접힌 구간 수
    """
    n, i, folded = len(values), 0, 0
    while i < n:
        j = i
        while (j + 1 < n and values[j + 1] == values[i] and values[i] is not None
               and (keys is None or keys[j + 1] == keys[i])):
            j += 1
        if j > i:
            folded += 1
            if mode == "merge":
                for k in range(i + 1, j + 1):
                    ws.cell(row=start + k, column=col).value = None
                ws.merge_cells(start_row=start + i, start_column=col,
                               end_row=start + j, end_column=col)
            else:
                for k in range(i + 1, j + 1):
                    c = ws.cell(row=start + k, column=col)
                    c.value, c.font, c.border, c.alignment = "〃", BODY, BOX, CEN
        i = j + 1
    return folded


def col_of(headers, name):
    return headers.index(name) + 1
